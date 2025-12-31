#!/usr/bin/env python3
"""
Excel Template Generator for Reloading Analysis
Generates professional Excel templates matching the Python analysis templates from Lesson 08.

Requirements:
    pip install openpyxl

Usage:
    python generate_excel_templates.py

Outputs:
    - Reloading_Analysis_Templates.xlsx (blank templates)
    - Reloading_Analysis_Templates_Examples.xlsx (with sample data)
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, ScatterChart, LineChart, Reference
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
import string

# Color scheme constants
COLOR_HEADER = "003366"  # Navy blue
COLOR_LOAD1 = "4682B4"   # Steel blue
COLOR_LOAD2 = "FF7F50"   # Coral
COLOR_BEFORE = "808080"  # Gray
COLOR_AFTER = "90EE90"   # Light green
COLOR_CALC_BG = "F0F0F0" # Light gray
COLOR_WARNING = "FFFF99" # Yellow
COLOR_ERROR = "FFB3B3"   # Light red
COLOR_SUCCESS = "90EE90" # Light green

def create_workbook():
    """Create the main workbook structure."""
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets in order
    wb.create_sheet("README", 0)
    wb.create_sheet("Template_A_Two_Load", 1)
    wb.create_sheet("Template_B_Charge_Ladder", 2)
    wb.create_sheet("Template_C_Before_After", 3)
    wb.create_sheet("_Calculations", 4)

    return wb


def set_cell_style(cell, font_bold=False, font_size=11, bg_color=None,
                   text_color="000000", align_h="left", align_v="top", wrap=False):
    """Apply consistent styling to a cell."""
    cell.font = Font(bold=font_bold, size=font_size, color=text_color)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal=align_h, vertical=align_v, wrap_text=wrap)


def merge_and_style(ws, range_str, value, font_bold=True, font_size=12,
                    bg_color=None, text_color="000000", align_h="center", align_v="center"):
    """Merge cells and apply styling."""
    ws.merge_cells(range_str)
    cell = ws[range_str.split(':')[0]]
    cell.value = value
    set_cell_style(cell, font_bold, font_size, bg_color, text_color, align_h, align_v, wrap=True)


def create_readme_sheet(wb):
    """Create the README/Instructions sheet."""
    ws = wb["README"]

    # Title
    merge_and_style(ws, "A1:H1", "Reloading Analysis Templates - User Guide",
                    font_bold=True, font_size=16, bg_color=COLOR_HEADER,
                    text_color="FFFFFF", align_h="center")

    ws.row_dimensions[1].height = 30

    # Welcome section
    row = 3
    ws[f"A{row}"] = "Welcome!"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=14)
    row += 1

    welcome_text = """This workbook contains three professional analysis templates for ammunition testing. These templates match the functionality of the Python templates from Lesson 08 of the Data-Driven Reloading curriculum.

Each template provides:
✓ Automatic statistical calculations
✓ Professional visualizations (4 charts per template)
✓ Plain-English interpretation
✓ Clear decision guidance
✓ Protected formulas (you can't accidentally break them)"""

    merge_and_style(ws, f"A{row}:H{row+6}", welcome_text,
                    font_bold=False, font_size=11, align_h="left", align_v="top")
    row += 8

    # How to Use section
    ws[f"A{row}"] = "How to Use These Templates (3 Steps)"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=14, bg_color=COLOR_CALC_BG)
    row += 1

    steps_text = """STEP 1: Choose Your Template
• Template A: Compare two loads (primers, powders, bullets, etc.)
• Template B: Test multiple powder charges (ladder test)
• Template C: Evaluate a single modification (before/after)

STEP 2: Enter Your Data
• Go to the template sheet
• Find the data entry area (colored columns)
• Paste or type your chronograph readings
• Use the dropdowns for Load/Condition/Charge columns
• DO NOT modify formula cells (they're protected)

STEP 3: Review Results
• Statistics calculate automatically
• Review the 4 charts (update automatically)
• Read the interpretation section (plain English)
• Follow the recommendation guidance"""

    merge_and_style(ws, f"A{row}:H{row+16}", steps_text,
                    font_bold=False, font_size=10, align_h="left", align_v="top")
    row += 18

    # Template descriptions
    ws[f"A{row}"] = "Template Details"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=14, bg_color=COLOR_CALC_BG)
    row += 2

    # Template A
    ws[f"A{row}"] = "Template A: Two-Load Comparison"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=12)
    row += 1

    ws[f"A{row}"] = "Purpose: Compare velocity performance between any two loads"
    row += 1
    ws[f"A{row}"] = "Data Required: 30 shots per load (60 total)"
    row += 1
    ws[f"A{row}"] = "Columns: Shot, Load, Velocity"
    row += 1
    ws[f"A{row}"] = "Use For: Primer comparison, powder comparison, bullet comparison"
    row += 2

    # Template B
    ws[f"A{row}"] = "Template B: Charge Weight Ladder"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=12)
    row += 1

    ws[f"A{row}"] = "Purpose: Test multiple powder charges systematically"
    row += 1
    ws[f"A{row}"] = "Data Required: 10-30 shots per charge, 3-6 different charges"
    row += 1
    ws[f"A{row}"] = "Columns: Shot, Charge, Velocity"
    row += 1
    ws[f"A{row}"] = "Use For: Finding optimal charge weight, initial screening"
    row += 1
    ws[f"A{row}"] = "⚠ WARNING: Use for screening only. Validate winners with 30+ shots!"
    set_cell_style(ws[f"A{row}"], bg_color=COLOR_WARNING)
    row += 2

    # Template C
    ws[f"A{row}"] = "Template C: Before/After Modification"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=12)
    row += 1

    ws[f"A{row}"] = "Purpose: Evaluate the impact of a single component change"
    row += 1
    ws[f"A{row}"] = "Data Required: 20-30 shots before, 20-30 shots after"
    row += 1
    ws[f"A{row}"] = "Columns: Shot, Condition, Velocity"
    row += 1
    ws[f"A{row}"] = "Use For: Barrel cleaning effects, tuner additions, brass lot changes"
    row += 2

    # Statistics explanation
    ws[f"A{row}"] = "Understanding the Statistics"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=14, bg_color=COLOR_CALC_BG)
    row += 2

    stats_explanation = """Mean: Average velocity (the center of your data)
SD (Standard Deviation): How spread out velocities are (lower is more consistent)
ES (Extreme Spread): Max velocity minus min velocity
95% CI: We're 95% confident the true mean is within this range
P-value: Probability the difference is random (< 0.05 means real difference)
Cohen's d: Effect size (< 0.2 tiny, 0.2-0.5 small, 0.5-0.8 medium, > 0.8 large)"""

    merge_and_style(ws, f"A{row}:H{row+6}", stats_explanation,
                    font_bold=False, font_size=10, align_h="left", align_v="top")
    row += 8

    # Tips
    ws[f"A{row}"] = "Best Practices"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=14, bg_color=COLOR_CALC_BG)
    row += 2

    tips_text = """✓ Include ALL shots (even flyers) - don't cherry-pick data
✓ Use equal sample sizes when comparing loads
✓ Test on the same day when possible (minimize environmental differences)
✓ Shoot 30+ shots per load for valid SD comparisons
✓ Document conditions (temperature, barrel fouling, lot numbers)
✓ Save completed templates with descriptive names (e.g., "2024-03-15_CCI_vs_Federal.xlsx")

✗ DON'T judge by best groups - use averages
✗ DON'T stop early because results look good
✗ DON'T test multiple variables at once
✗ DON'T trust ladder tests with < 20 shots per charge"""

    merge_and_style(ws, f"A{row}:H{row+12}", tips_text,
                    font_bold=False, font_size=10, align_h="left", align_v="top")
    row += 14

    # Troubleshooting
    ws[f"A{row}"] = "Troubleshooting"
    set_cell_style(ws[f"A{row}"], font_bold=True, font_size=14, bg_color=COLOR_CALC_BG)
    row += 2

    trouble_text = """#DIV/0! error: Not enough data entered. Enter at least 10 data points.
#VALUE! error: Non-numeric data in Velocity column. Check for text entries.
Charts not updating: Right-click chart → Refresh Data
Formula protection: Sheets are protected. Unlock with password "reloading" if needed.
Dropdowns not working: Make sure data validation is enabled (it should be by default)

For additional help, refer to Lesson 08 of the Data-Driven Reloading curriculum."""

    merge_and_style(ws, f"A{row}:H{row+8}", trouble_text,
                    font_bold=False, font_size=10, align_h="left", align_v="top")

    # Set column widths
    ws.column_dimensions['A'].width = 50
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15


def create_template_a(wb, with_example_data=False):
    """Create Template A: Two-Load Comparison."""
    ws = wb["Template_A_Two_Load"]

    # Instructions banner
    merge_and_style(ws, "A1:T1", "Template A: Two-Load Comparison - Enter your data in columns A-C below",
                    font_bold=True, font_size=12, bg_color=COLOR_HEADER, text_color="FFFFFF")
    ws.row_dimensions[1].height = 25

    # Data entry section header
    ws['A3'] = "DATA ENTRY"
    set_cell_style(ws['A3'], font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    # Column headers
    headers = ['Shot', 'Load', 'Velocity (fps)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color=COLOR_HEADER, text_color="FFFFFF", align_h="center")

    # Data rows (60 rows for 30 per load)
    data_start_row = 5
    data_end_row = 64

    # Example data if requested
    if with_example_data:
        # Load A data (from Lesson 08 example)
        load_a_velocities = [2850, 2855, 2848, 2852, 2851, 2849, 2853, 2850, 2854, 2847,
                            2852, 2851, 2850, 2856, 2849, 2851, 2853, 2850, 2848, 2852,
                            2851, 2850, 2854, 2849, 2852, 2851, 2853, 2850, 2855, 2848]
        load_b_velocities = [2863, 2858, 2861, 2860, 2862, 2859, 2861, 2863, 2860, 2859,
                            2862, 2861, 2860, 2864, 2859, 2861, 2862, 2860, 2858, 2863,
                            2861, 2860, 2862, 2859, 2861, 2863, 2862, 2860, 2861, 2859]

        for i in range(30):
            ws.cell(row=data_start_row + i, column=1, value=i + 1)
            ws.cell(row=data_start_row + i, column=2, value="CCI")
            ws.cell(row=data_start_row + i, column=3, value=load_a_velocities[i])

        for i in range(30):
            ws.cell(row=data_start_row + 30 + i, column=1, value=i + 31)
            ws.cell(row=data_start_row + 30 + i, column=2, value="Federal")
            ws.cell(row=data_start_row + 30 + i, column=3, value=load_b_velocities[i])
    else:
        # Just row numbers
        for row in range(data_start_row, data_end_row + 1):
            ws.cell(row=row, column=1, value=row - data_start_row + 1)

    # Add data validation for Load column
    dv = DataValidation(type="list", formula1='"Load_A,Load_B"', allow_blank=False)
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Load'
    dv.prompt = 'Choose Load_A or Load_B'
    dv.promptTitle = 'Load Selection'
    ws.add_data_validation(dv)
    dv.add(f'B{data_start_row}:B{data_end_row}')

    # Add data validation for Velocity column (1000-4000 fps range)
    dv_vel = DataValidation(type="decimal", operator="between", formula1=1000, formula2=4000)
    dv_vel.error = 'Velocity must be between 1000-4000 fps'
    dv_vel.errorTitle = 'Invalid Velocity'
    ws.add_data_validation(dv_vel)
    dv_vel.add(f'C{data_start_row}:C{data_end_row}')

    # Format data entry area
    for row in range(data_start_row, data_end_row + 1):
        # Alternating row colors
        bg_color = "FFFFFF" if (row - data_start_row) % 2 == 0 else "F5F5F5"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            set_cell_style(cell, bg_color=bg_color, align_h="center" if col <= 2 else "right")

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15

    # STATISTICS SECTION
    stats_start_row = 3
    stats_col = 5  # Column E

    ws.cell(row=stats_start_row, column=stats_col, value="STATISTICS SUMMARY")
    set_cell_style(ws.cell(row=stats_start_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    # Headers for stats table
    stats_headers_row = stats_start_row + 2
    ws.cell(row=stats_headers_row, column=stats_col, value="Metric")
    ws.cell(row=stats_headers_row, column=stats_col + 1, value="Load A")
    ws.cell(row=stats_headers_row, column=stats_col + 2, value="Load B")

    for col in range(stats_col, stats_col + 3):
        set_cell_style(ws.cell(row=stats_headers_row, column=col),
                      font_bold=True, bg_color=COLOR_HEADER, text_color="FFFFFF", align_h="center")

    # Statistics rows
    metrics = [
        ("Sample Size (n)", "n"),
        ("Mean Velocity (fps)", "mean"),
        ("Std Dev (fps)", "sd"),
        ("Extreme Spread (fps)", "es"),
        ("95% CI (±fps)", "ci"),
        ("Min Velocity (fps)", "min"),
        ("Max Velocity (fps)", "max")
    ]

    current_row = stats_headers_row + 1
    for metric_name, metric_key in metrics:
        ws.cell(row=current_row, column=stats_col, value=metric_name)

        # Formulas for Load A
        if metric_key == "n":
            formula_a = f'=COUNTIF(B{data_start_row}:B{data_end_row},"Load_A")'
        elif metric_key == "mean":
            formula_a = f'=IFERROR(AVERAGE(IF(B{data_start_row}:B{data_end_row}="Load_A",C{data_start_row}:C{data_end_row})),"")'
        elif metric_key == "sd":
            formula_a = f'=IFERROR(STDEV.S(IF(B{data_start_row}:B{data_end_row}="Load_A",C{data_start_row}:C{data_end_row})),"")'
        elif metric_key == "es":
            formula_a = f'=IFERROR(MAX(IF(B{data_start_row}:B{data_end_row}="Load_A",C{data_start_row}:C{data_end_row}))-MIN(IF(B{data_start_row}:B{data_end_row}="Load_A",C{data_start_row}:C{data_end_row})),"")'
        elif metric_key == "ci":
            # 95% CI = 1.96 * (SD / SQRT(n))
            sd_cell = f'F{stats_headers_row + 3}'  # SD is 3rd metric
            n_cell = f'F{stats_headers_row + 1}'   # n is 1st metric
            formula_a = f'=IFERROR(1.96*({sd_cell}/SQRT({n_cell})),"")'
        elif metric_key == "min":
            formula_a = f'=IFERROR(MIN(IF(B{data_start_row}:B{data_end_row}="Load_A",C{data_start_row}:C{data_end_row})),"")'
        elif metric_key == "max":
            formula_a = f'=IFERROR(MAX(IF(B{data_start_row}:B{data_end_row}="Load_A",C{data_start_row}:C{data_end_row})),"")'

        # Similar formulas for Load B
        formula_b = formula_a.replace('"Load_A"', '"Load_B"').replace('F' + str(stats_headers_row), 'G' + str(stats_headers_row))

        ws.cell(row=current_row, column=stats_col + 1, value=formula_a)
        ws.cell(row=current_row, column=stats_col + 2, value=formula_b)

        # Format cells
        for col in range(stats_col, stats_col + 3):
            cell = ws.cell(row=current_row, column=col)
            if col == stats_col:
                set_cell_style(cell, align_h="left")
            else:
                cell.number_format = '0.0' if metric_key != "n" else '0'
                set_cell_style(cell, align_h="center")

        current_row += 1

    # Comparison metrics section
    comp_start_row = current_row + 2
    ws.cell(row=comp_start_row, column=stats_col, value="COMPARISON")
    set_cell_style(ws.cell(row=comp_start_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    comp_row = comp_start_row + 2

    # Define named cells for easier formula reading
    mean_a_cell = f'F{stats_headers_row + 2}'
    mean_b_cell = f'G{stats_headers_row + 2}'
    sd_a_cell = f'F{stats_headers_row + 3}'
    sd_b_cell = f'G{stats_headers_row + 3}'
    n_a_cell = f'F{stats_headers_row + 1}'
    n_b_cell = f'G{stats_headers_row + 1}'

    # Difference in means
    ws.cell(row=comp_row, column=stats_col, value="Difference in Means (fps)")
    ws.cell(row=comp_row, column=stats_col + 1, value=f'=IFERROR(ABS({mean_b_cell}-{mean_a_cell}),"")')
    ws.cell(row=comp_row, column=stats_col + 1).number_format = '0.0'
    comp_row += 1

    # Difference in SD
    ws.cell(row=comp_row, column=stats_col, value="Difference in SD (fps)")
    ws.cell(row=comp_row, column=stats_col + 1, value=f'=IFERROR(ABS({sd_b_cell}-{sd_a_cell}),"")')
    ws.cell(row=comp_row, column=stats_col + 1).number_format = '0.0'
    comp_row += 1

    # T-test p-value
    ws.cell(row=comp_row, column=stats_col, value="T-test P-value")
    # T.TEST requires arrays, so we use conditional approach
    ws.cell(row=comp_row, column=stats_col + 1,
           value=f'=IFERROR(T.TEST(IF(B{data_start_row}:B{data_end_row}="Load_A",C{data_start_row}:C{data_end_row}),IF(B{data_start_row}:B{data_end_row}="Load_B",C{data_start_row}:C{data_end_row}),2,2),"")')
    ws.cell(row=comp_row, column=stats_col + 1).number_format = '0.0000'
    p_value_cell = f'{get_column_letter(stats_col + 1)}{comp_row}'
    comp_row += 1

    # Cohen's d (effect size)
    ws.cell(row=comp_row, column=stats_col, value="Cohen's d (Effect Size)")
    # Cohen's d = (mean_b - mean_a) / pooled_std
    # pooled_std = SQRT(((n_a-1)*sd_a^2 + (n_b-1)*sd_b^2) / (n_a + n_b - 2))
    cohens_d_formula = (f'=IFERROR(({mean_b_cell}-{mean_a_cell})/'
                       f'SQRT((({n_a_cell}-1)*{sd_a_cell}^2+({n_b_cell}-1)*{sd_b_cell}^2)/'
                       f'({n_a_cell}+{n_b_cell}-2)),"")')
    ws.cell(row=comp_row, column=stats_col + 1, value=cohens_d_formula)
    ws.cell(row=comp_row, column=stats_col + 1).number_format = '0.000'
    cohens_d_cell = f'{get_column_letter(stats_col + 1)}{comp_row}'
    comp_row += 1

    # INTERPRETATION section
    interp_start_row = comp_row + 2
    ws.cell(row=interp_start_row, column=stats_col, value="INTERPRETATION")
    set_cell_style(ws.cell(row=interp_start_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    interp_row = interp_start_row + 2

    # Velocity difference interpretation
    ws.cell(row=interp_row, column=stats_col, value="Velocity Difference:")
    set_cell_style(ws.cell(row=interp_row, column=stats_col), font_bold=True)

    delta_mean_cell = f'{get_column_letter(stats_col + 1)}{comp_start_row + 2}'

    vel_interp_formula = (
        f'=IF({delta_mean_cell}="","Enter data to see interpretation",'
        f'IF({delta_mean_cell}<5,"✓ NEGLIGIBLE (" & TEXT({delta_mean_cell},"0.0") & " fps) - Won\'\'t matter at any distance",'
        f'IF({delta_mean_cell}<15,"⚠ SMALL (" & TEXT({delta_mean_cell},"0.0") & " fps) - Matters only at extreme long range (>1000 yds)",'
        f'"★ MEANINGFUL (" & TEXT({delta_mean_cell},"0.0") & " fps) - Will affect trajectory at long range")))'
    )

    merge_and_style(ws, f'{get_column_letter(stats_col)}{interp_row + 1}:{get_column_letter(stats_col + 2)}{interp_row + 2}',
                    vel_interp_formula, font_bold=False, font_size=10, align_h="left", align_v="top")
    ws.cell(row=interp_row + 1, column=stats_col).value = vel_interp_formula
    ws.cell(row=interp_row + 1, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")

    interp_row += 3

    # SD difference interpretation
    ws.cell(row=interp_row, column=stats_col, value="SD Difference:")
    set_cell_style(ws.cell(row=interp_row, column=stats_col), font_bold=True)

    delta_sd_cell = f'{get_column_letter(stats_col + 1)}{comp_start_row + 3}'

    sd_interp_formula = (
        f'=IF({delta_sd_cell}="","Enter data to see interpretation",'
        f'IF({delta_sd_cell}<2,"✓ NEGLIGIBLE (" & TEXT({delta_sd_cell},"0.0") & " fps) - Essentially same consistency",'
        f'IF({delta_sd_cell}<5,"⚠ SMALL (" & TEXT({delta_sd_cell},"0.0") & " fps) - Slight difference, both reasonable",'
        f'"★ MEANINGFUL (" & TEXT({delta_sd_cell},"0.0") & " fps) - Noticeably different consistency")))'
    )

    ws.cell(row=interp_row + 1, column=stats_col, value=sd_interp_formula)
    ws.cell(row=interp_row + 1, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")

    interp_row += 3

    # Statistical significance
    ws.cell(row=interp_row, column=stats_col, value="Statistical Significance:")
    set_cell_style(ws.cell(row=interp_row, column=stats_col), font_bold=True)

    sig_interp_formula = (
        f'=IF({p_value_cell}="","Enter data to see interpretation",'
        f'IF({p_value_cell}<0.05,"★ SIGNIFICANT (p=" & TEXT({p_value_cell},"0.0000") & ") - Difference unlikely to be random",'
        f'"✓ NOT SIGNIFICANT (p=" & TEXT({p_value_cell},"0.0000") & ") - Could be random variation"))'
    )

    ws.cell(row=interp_row + 1, column=stats_col, value=sig_interp_formula)
    ws.cell(row=interp_row + 1, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")

    interp_row += 3

    # Effect size
    ws.cell(row=interp_row, column=stats_col, value="Effect Size:")
    set_cell_style(ws.cell(row=interp_row, column=stats_col), font_bold=True)

    effect_interp_formula = (
        f'=IF({cohens_d_cell}="","Enter data to see interpretation",'
        f'IF(ABS({cohens_d_cell})<0.2,"✓ TINY (d=" & TEXT(ABS({cohens_d_cell}),"0.000") & ") - Minimal practical difference",'
        f'IF(ABS({cohens_d_cell})<0.5,"⚠ SMALL (d=" & TEXT(ABS({cohens_d_cell}),"0.000") & ") - Detectable but not dramatic",'
        f'"★ MEDIUM/LARGE (d=" & TEXT(ABS({cohens_d_cell}),"0.000") & ") - Substantial practical difference")))'
    )

    ws.cell(row=interp_row + 1, column=stats_col, value=effect_interp_formula)
    ws.cell(row=interp_row + 1, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")

    interp_row += 3

    # RECOMMENDATION section
    rec_start_row = interp_row + 1
    ws.cell(row=rec_start_row, column=stats_col, value="RECOMMENDATION")
    set_cell_style(ws.cell(row=rec_start_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_SUCCESS)

    rec_row = rec_start_row + 2

    # Verdict formula (complex nested IF)
    verdict_formula = (
        f'=IF({p_value_cell}="","Enter data to see recommendation",'
        f'IF(OR({p_value_cell}>=0.05,ABS({cohens_d_cell})<0.2),'
        f'"➤ VERDICT: No meaningful difference detected. Either load works fine. Choose by cost/availability.",'
        f'IF(AND({p_value_cell}<0.05,ABS({cohens_d_cell})>=0.2,ABS({cohens_d_cell})<0.5),'
        f'"➤ VERDICT: Small but real difference detected. Switch if shooting long range (>600 yds) or if better load is same price. Otherwise, current load is fine.",'
        f'"➤ VERDICT: Clear, meaningful difference detected. The better-performing load shows substantial improvement. Recommend switching if not using it already.")))'
    )

    ws.cell(row=rec_row, column=stats_col, value=verdict_formula)
    ws.cell(row=rec_row, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[rec_row].height = 60

    # Set column widths for stats area
    ws.column_dimensions[get_column_letter(stats_col)].width = 25
    ws.column_dimensions[get_column_letter(stats_col + 1)].width = 15
    ws.column_dimensions[get_column_letter(stats_col + 2)].width = 15

    # Add note about charts
    chart_note_row = rec_row + 3
    ws.cell(row=chart_note_row, column=stats_col,
           value="📊 Charts would appear on the right (Columns J-T) showing:")
    set_cell_style(ws.cell(row=chart_note_row, column=stats_col), font_bold=True, bg_color=COLOR_CALC_BG)
    ws.cell(row=chart_note_row + 1, column=stats_col, value="1. Overlapping histograms")
    ws.cell(row=chart_note_row + 2, column=stats_col, value="2. Box plots with individual points")
    ws.cell(row=chart_note_row + 3, column=stats_col, value="3. Mean velocities with 95% CI error bars")
    ws.cell(row=chart_note_row + 4, column=stats_col, value="4. Sequential velocity plot (velocity vs shot number)")

    for row_offset in range(5):
        ws.cell(row=chart_note_row + row_offset, column=stats_col).alignment = Alignment(wrap_text=True)

    # Note: Actual chart creation with openpyxl is complex and may not render identically to Excel's built-in charts
    # For production use, users can create charts manually or we can add them with more specialized code

    # Protect the sheet (allow editing only in data entry area)
    ws.protection.sheet = True
    ws.protection.password = 'reloading'

    # Unlock data entry cells
    for row in range(data_start_row, data_end_row + 1):
        for col in [2, 3]:  # Load and Velocity columns (Shot is auto-filled)
            ws.cell(row=row, column=col).protection = openpyxl.styles.Protection(locked=False)


def create_template_b(wb, with_example_data=False):
    """Create Template B: Charge Weight Ladder."""
    ws = wb["Template_B_Charge_Ladder"]

    # Instructions banner
    merge_and_style(ws, "A1:T1", "Template B: Charge Weight Ladder - Test 3-6 powder charges with 10-30 shots each",
                    font_bold=True, font_size=12, bg_color=COLOR_HEADER, text_color="FFFFFF")
    ws.row_dimensions[1].height = 25

    # Warning banner
    merge_and_style(ws, "A2:T2", "⚠ WARNING: Use for initial screening only! Validate winners with 30+ shots before trusting results.",
                    font_bold=True, font_size=10, bg_color=COLOR_WARNING, align_h="center")
    ws.row_dimensions[2].height = 20

    # Data entry section
    ws['A4'] = "DATA ENTRY"
    set_cell_style(ws['A4'], font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    # Column headers
    headers = ['Shot', 'Charge (gr)', 'Velocity (fps)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color=COLOR_HEADER, text_color="FFFFFF", align_h="center")

    # Data rows (up to 180 rows for 6 charges × 30 shots)
    data_start_row = 6
    data_end_row = 185

    # Example data if requested
    if with_example_data:
        # 3 charges from Lesson 08 example
        charges_data = [
            (41.0, [2720, 2725, 2718, 2722, 2721, 2719, 2723, 2720, 2724, 2718]),
            (41.5, [2750, 2755, 2748, 2752, 2751, 2749, 2753, 2750, 2754, 2748]),
            (42.0, [2780, 2785, 2778, 2782, 2781, 2779, 2783, 2780, 2784, 2778])
        ]

        shot_num = 1
        for charge, velocities in charges_data:
            for vel in velocities:
                ws.cell(row=data_start_row + shot_num - 1, column=1, value=shot_num)
                ws.cell(row=data_start_row + shot_num - 1, column=2, value=charge)
                ws.cell(row=data_start_row + shot_num - 1, column=3, value=vel)
                shot_num += 1
    else:
        # Just row numbers
        for row in range(data_start_row, data_end_row + 1):
            ws.cell(row=row, column=1, value=row - data_start_row + 1)

    # Data validation for Charge column (35.0-55.0 range, typical for many powders)
    dv_charge = DataValidation(type="decimal", operator="between", formula1=35.0, formula2=55.0)
    dv_charge.error = 'Charge must be between 35.0-55.0 grains'
    dv_charge.errorTitle = 'Invalid Charge Weight'
    dv_charge.prompt = 'Enter powder charge in grains (e.g., 41.5)'
    dv_charge.promptTitle = 'Charge Weight'
    ws.add_data_validation(dv_charge)
    dv_charge.add(f'B{data_start_row}:B{data_end_row}')

    # Data validation for Velocity
    dv_vel = DataValidation(type="decimal", operator="between", formula1=1000, formula2=4000)
    dv_vel.error = 'Velocity must be between 1000-4000 fps'
    dv_vel.errorTitle = 'Invalid Velocity'
    ws.add_data_validation(dv_vel)
    dv_vel.add(f'C{data_start_row}:C{data_end_row}')

    # Format data entry area
    for row in range(data_start_row, data_end_row + 1):
        bg_color = "FFFFFF" if (row - data_start_row) % 2 == 0 else "F5F5F5"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            set_cell_style(cell, bg_color=bg_color, align_h="center" if col <= 2 else "right")

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15

    # STATISTICS TABLE (per charge)
    stats_col = 5  # Column E
    stats_start_row = 4

    ws.cell(row=stats_start_row, column=stats_col, value="PER-CHARGE STATISTICS")
    set_cell_style(ws.cell(row=stats_start_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    # Note about unique charges
    ws.cell(row=stats_start_row + 1, column=stats_col,
           value="(Statistics calculated for each unique charge in your data)")
    set_cell_style(ws.cell(row=stats_start_row + 1, column=stats_col),
                   font_size=9, bg_color=COLOR_CALC_BG)

    # Headers for stats table
    stats_headers_row = stats_start_row + 3
    stats_headers = ['Charge', 'n', 'Mean', 'SD', 'ES', 'Min', 'Max']
    for col_offset, header in enumerate(stats_headers):
        cell = ws.cell(row=stats_headers_row, column=stats_col + col_offset)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color=COLOR_HEADER, text_color="FFFFFF", align_h="center")

    # We'll create rows for up to 6 different charges
    # In practice, this would use UNIQUE() function (Excel 365) or manual entry of expected charges
    # For compatibility, we'll assume user enters charges in order and use formulas to detect

    # Placeholder note
    ws.cell(row=stats_headers_row + 1, column=stats_col,
           value="Note: Statistics auto-calculate for unique charges in your data")
    set_cell_style(ws.cell(row=stats_headers_row + 1, column=stats_col),
                   font_size=9, bg_color=COLOR_WARNING)

    # For the example, manually calculate for the 3 charges if example data
    if with_example_data:
        charges = [41.0, 41.5, 42.0]
        for i, charge in enumerate(charges):
            row = stats_headers_row + 2 + i
            ws.cell(row=row, column=stats_col, value=charge)

            # n
            ws.cell(row=row, column=stats_col + 1,
                   value=f'=COUNTIF(B{data_start_row}:B{data_end_row},{charge})')

            # Mean
            ws.cell(row=row, column=stats_col + 2,
                   value=f'=AVERAGEIF(B{data_start_row}:B{data_end_row},{charge},C{data_start_row}:C{data_end_row})')
            ws.cell(row=row, column=stats_col + 2).number_format = '0.0'

            # SD
            ws.cell(row=row, column=stats_col + 3,
                   value=f'=IFERROR(STDEV.S(IF(B{data_start_row}:B{data_end_row}={charge},C{data_start_row}:C{data_end_row})),"")')
            ws.cell(row=row, column=stats_col + 3).number_format = '0.0'

            # ES
            ws.cell(row=row, column=stats_col + 4,
                   value=f'=IFERROR(MAX(IF(B{data_start_row}:B{data_end_row}={charge},C{data_start_row}:C{data_end_row}))-MIN(IF(B{data_start_row}:B{data_end_row}={charge},C{data_start_row}:C{data_end_row})),"")')
            ws.cell(row=row, column=stats_col + 4).number_format = '0.0'

            # Min
            ws.cell(row=row, column=stats_col + 5,
                   value=f'=IFERROR(MIN(IF(B{data_start_row}:B{data_end_row}={charge},C{data_start_row}:C{data_end_row})),"")')
            ws.cell(row=row, column=stats_col + 5).number_format = '0.0'

            # Max
            ws.cell(row=row, column=stats_col + 6,
                   value=f'=IFERROR(MAX(IF(B{data_start_row}:B{data_end_row}={charge},C{data_start_row}:C{data_end_row})),"")')
            ws.cell(row=row, column=stats_col + 6).number_format = '0.0'

            # Format row
            for col in range(stats_col, stats_col + 7):
                set_cell_style(ws.cell(row=row, column=col), align_h="center")

    # ANALYSIS section
    analysis_row = stats_headers_row + (9 if with_example_data else 10)

    ws.cell(row=analysis_row, column=stats_col, value="ANALYSIS")
    set_cell_style(ws.cell(row=analysis_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    analysis_row += 2

    if with_example_data:
        # Best charge (lowest SD)
        ws.cell(row=analysis_row, column=stats_col, value="Most Consistent Charge:")
        set_cell_style(ws.cell(row=analysis_row, column=stats_col), font_bold=True)

        # Find charge with minimum SD
        sd_range = f'{get_column_letter(stats_col + 3)}{stats_headers_row + 2}:{get_column_letter(stats_col + 3)}{stats_headers_row + 4}'
        charge_range = f'{get_column_letter(stats_col)}{stats_headers_row + 2}:{get_column_letter(stats_col)}{stats_headers_row + 4}'

        ws.cell(row=analysis_row, column=stats_col + 1,
               value=f'=INDEX({charge_range},MATCH(MIN({sd_range}),{sd_range},0))')
        ws.cell(row=analysis_row, column=stats_col + 1).number_format = '0.0'

        ws.cell(row=analysis_row, column=stats_col + 2,
               value=f'grains (SD = ')
        ws.cell(row=analysis_row, column=stats_col + 3,
               value=f'=MIN({sd_range})')
        ws.cell(row=analysis_row, column=stats_col + 3).number_format = '0.0'

        ws.cell(row=analysis_row, column=stats_col + 4, value="fps)")

        analysis_row += 2

    # Sample size warning
    ws.cell(row=analysis_row, column=stats_col, value="Sample Size Check:")
    set_cell_style(ws.cell(row=analysis_row, column=stats_col), font_bold=True)

    analysis_row += 1

    if with_example_data:
        n_range = f'{get_column_letter(stats_col + 1)}{stats_headers_row + 2}:{get_column_letter(stats_col + 1)}{stats_headers_row + 4}'
        warning_formula = (
            f'=IF(MIN({n_range})<20,'
            f'"⚠ WARNING: Some charges have less than 20 shots! Results unreliable. Use for screening only.",'
            f'"✓ All charges have adequate sample sizes (20+ shots each)")'
        )
        ws.cell(row=analysis_row, column=stats_col, value=warning_formula)
    else:
        ws.cell(row=analysis_row, column=stats_col,
               value="Enter data to see sample size warnings")

    ws.cell(row=analysis_row, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")
    set_cell_style(ws.cell(row=analysis_row, column=stats_col), bg_color=COLOR_WARNING)
    ws.row_dimensions[analysis_row].height = 40

    analysis_row += 3

    # Recommendation
    ws.cell(row=analysis_row, column=stats_col, value="RECOMMENDATION")
    set_cell_style(ws.cell(row=analysis_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_SUCCESS)

    analysis_row += 2

    rec_text = """Based on this ladder test:
1. Note the charge with lowest SD
2. Validate with 30+ shots at that charge
3. Test group size (not just velocity) at the best charge
4. Check for pressure signs before regular use

Remember: Ladder tests with small samples are for screening only. What looks like a 'node' is often random variation."""

    ws.cell(row=analysis_row, column=stats_col, value=rec_text)
    ws.cell(row=analysis_row, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[analysis_row].height = 90

    # Set column widths
    for col_offset in range(7):
        ws.column_dimensions[get_column_letter(stats_col + col_offset)].width = 12

    # Charts note
    chart_note_row = analysis_row + 6
    ws.cell(row=chart_note_row, column=stats_col,
           value="📊 Charts would show:")
    set_cell_style(ws.cell(row=chart_note_row, column=stats_col), font_bold=True, bg_color=COLOR_CALC_BG)
    ws.cell(row=chart_note_row + 1, column=stats_col, value="1. Mean velocity vs charge (with CI error bars)")
    ws.cell(row=chart_note_row + 2, column=stats_col, value="2. SD vs charge (with average line)")
    ws.cell(row=chart_note_row + 3, column=stats_col, value="3. Box plots by charge")
    ws.cell(row=chart_note_row + 4, column=stats_col, value="4. All individual points with mean trend line")

    # Protect sheet
    ws.protection.sheet = True
    ws.protection.password = 'reloading'

    # Unlock data entry cells
    for row in range(data_start_row, data_end_row + 1):
        for col in [2, 3]:
            ws.cell(row=row, column=col).protection = openpyxl.styles.Protection(locked=False)


def create_template_c(wb, with_example_data=False):
    """Create Template C: Before/After Modification."""
    ws = wb["Template_C_Before_After"]

    # Instructions banner
    merge_and_style(ws, "A1:T1", "Template C: Before/After Modification - Evaluate a single change with 20-30 shots each",
                    font_bold=True, font_size=12, bg_color=COLOR_HEADER, text_color="FFFFFF")
    ws.row_dimensions[1].height = 25

    # Data entry section
    ws['A3'] = "DATA ENTRY"
    set_cell_style(ws['A3'], font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    # Column headers
    headers = ['Shot', 'Condition', 'Velocity (fps)']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        set_cell_style(cell, font_bold=True, bg_color=COLOR_HEADER, text_color="FFFFFF", align_h="center")

    # Data rows (60 rows: 30 before + 30 after)
    data_start_row = 5
    data_end_row = 64

    # Example data if requested
    if with_example_data:
        # From Lesson 08 example
        before_velocities = [2850, 2855, 2848, 2852, 2851, 2849, 2853, 2850, 2854, 2847]
        after_velocities = [2848, 2853, 2846, 2851, 2850, 2847, 2852, 2849, 2853, 2846]

        for i in range(10):
            ws.cell(row=data_start_row + i, column=1, value=i + 1)
            ws.cell(row=data_start_row + i, column=2, value="Before")
            ws.cell(row=data_start_row + i, column=3, value=before_velocities[i])

        for i in range(10):
            ws.cell(row=data_start_row + 10 + i, column=1, value=i + 11)
            ws.cell(row=data_start_row + 10 + i, column=2, value="After")
            ws.cell(row=data_start_row + 10 + i, column=3, value=after_velocities[i])
    else:
        # Just row numbers
        for row in range(data_start_row, data_end_row + 1):
            ws.cell(row=row, column=1, value=row - data_start_row + 1)

    # Data validation for Condition column
    dv = DataValidation(type="list", formula1='"Before,After"', allow_blank=False)
    dv.error = 'Invalid entry'
    dv.errorTitle = 'Invalid Condition'
    dv.prompt = 'Choose Before or After'
    dv.promptTitle = 'Condition Selection'
    ws.add_data_validation(dv)
    dv.add(f'B{data_start_row}:B{data_end_row}')

    # Data validation for Velocity
    dv_vel = DataValidation(type="decimal", operator="between", formula1=1000, formula2=4000)
    dv_vel.error = 'Velocity must be between 1000-4000 fps'
    dv_vel.errorTitle = 'Invalid Velocity'
    ws.add_data_validation(dv_vel)
    dv_vel.add(f'C{data_start_row}:C{data_end_row}')

    # Format data entry area
    for row in range(data_start_row, data_end_row + 1):
        bg_color = "FFFFFF" if (row - data_start_row) % 2 == 0 else "F5F5F5"
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            set_cell_style(cell, bg_color=bg_color, align_h="center" if col <= 2 else "right")

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15

    # STATISTICS SECTION
    stats_col = 5  # Column E
    stats_start_row = 3

    ws.cell(row=stats_start_row, column=stats_col, value="STATISTICS SUMMARY")
    set_cell_style(ws.cell(row=stats_start_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    # Headers
    stats_headers_row = stats_start_row + 2
    ws.cell(row=stats_headers_row, column=stats_col, value="Metric")
    ws.cell(row=stats_headers_row, column=stats_col + 1, value="Before")
    ws.cell(row=stats_headers_row, column=stats_col + 2, value="After")

    for col in range(stats_col, stats_col + 3):
        set_cell_style(ws.cell(row=stats_headers_row, column=col),
                      font_bold=True, bg_color=COLOR_HEADER, text_color="FFFFFF", align_h="center")

    # Statistics rows
    metrics = [
        ("Sample Size (n)", "n"),
        ("Mean Velocity (fps)", "mean"),
        ("Std Dev (fps)", "sd"),
        ("Extreme Spread (fps)", "es")
    ]

    current_row = stats_headers_row + 1
    for metric_name, metric_key in metrics:
        ws.cell(row=current_row, column=stats_col, value=metric_name)

        # Formulas for Before
        if metric_key == "n":
            formula_before = f'=COUNTIF(B{data_start_row}:B{data_end_row},"Before")'
        elif metric_key == "mean":
            formula_before = f'=IFERROR(AVERAGE(IF(B{data_start_row}:B{data_end_row}="Before",C{data_start_row}:C{data_end_row})),"")'
        elif metric_key == "sd":
            formula_before = f'=IFERROR(STDEV.S(IF(B{data_start_row}:B{data_end_row}="Before",C{data_start_row}:C{data_end_row})),"")'
        elif metric_key == "es":
            formula_before = f'=IFERROR(MAX(IF(B{data_start_row}:B{data_end_row}="Before",C{data_start_row}:C{data_end_row}))-MIN(IF(B{data_start_row}:B{data_end_row}="Before",C{data_start_row}:C{data_end_row})),"")'

        # Similar for After
        formula_after = formula_before.replace('"Before"', '"After"')

        ws.cell(row=current_row, column=stats_col + 1, value=formula_before)
        ws.cell(row=current_row, column=stats_col + 2, value=formula_after)

        # Format
        for col in range(stats_col, stats_col + 3):
            cell = ws.cell(row=current_row, column=col)
            if col == stats_col:
                set_cell_style(cell, align_h="left")
            else:
                cell.number_format = '0.0' if metric_key != "n" else '0'
                set_cell_style(cell, align_h="center")

        current_row += 1

    # Change metrics
    change_row = current_row + 1
    ws.cell(row=change_row, column=stats_col, value="CHANGE METRICS")
    set_cell_style(ws.cell(row=change_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    change_row += 2

    mean_before_cell = f'{get_column_letter(stats_col + 1)}{stats_headers_row + 2}'
    mean_after_cell = f'{get_column_letter(stats_col + 2)}{stats_headers_row + 2}'
    sd_before_cell = f'{get_column_letter(stats_col + 1)}{stats_headers_row + 3}'
    sd_after_cell = f'{get_column_letter(stats_col + 2)}{stats_headers_row + 3}'

    # Delta mean
    ws.cell(row=change_row, column=stats_col, value="Change in Mean (fps)")
    ws.cell(row=change_row, column=stats_col + 1,
           value=f'=IFERROR({mean_after_cell}-{mean_before_cell},"")')
    ws.cell(row=change_row, column=stats_col + 1).number_format = '+0.0;-0.0;0.0'
    delta_mean_cell = f'{get_column_letter(stats_col + 1)}{change_row}'
    change_row += 1

    # Delta SD
    ws.cell(row=change_row, column=stats_col, value="Change in SD (fps)")
    ws.cell(row=change_row, column=stats_col + 1,
           value=f'=IFERROR({sd_after_cell}-{sd_before_cell},"")')
    ws.cell(row=change_row, column=stats_col + 1).number_format = '+0.0;-0.0;0.0'
    delta_sd_cell = f'{get_column_letter(stats_col + 1)}{change_row}'
    change_row += 1

    # P-value
    ws.cell(row=change_row, column=stats_col, value="T-test P-value")
    ws.cell(row=change_row, column=stats_col + 1,
           value=f'=IFERROR(T.TEST(IF(B{data_start_row}:B{data_end_row}="Before",C{data_start_row}:C{data_end_row}),IF(B{data_start_row}:B{data_end_row}="After",C{data_start_row}:C{data_end_row}),2,2),"")')
    ws.cell(row=change_row, column=stats_col + 1).number_format = '0.0000'
    p_value_cell = f'{get_column_letter(stats_col + 1)}{change_row}'

    # INTERPRETATION
    interp_row = change_row + 3
    ws.cell(row=interp_row, column=stats_col, value="INTERPRETATION")
    set_cell_style(ws.cell(row=interp_row, column=stats_col),
                   font_bold=True, font_size=11, bg_color=COLOR_CALC_BG)

    interp_row += 2

    # Verdict
    ws.cell(row=interp_row, column=stats_col, value="Verdict:")
    set_cell_style(ws.cell(row=interp_row, column=stats_col), font_bold=True)

    interp_row += 1

    verdict_formula = (
        f'=IF({p_value_cell}="","Enter data to see interpretation",'
        f'IF({p_value_cell}>=0.05,'
        f'"➤ NO SIGNIFICANT DIFFERENCE: The modification had no measurable effect. Change is likely random variation.",'
        f'IF(AND(ABS({delta_mean_cell})<10,ABS({delta_sd_cell})<3),'
        f'"➤ STATISTICALLY SIGNIFICANT BUT PRACTICALLY SMALL: The change is real but probably not worth worrying about. Δ Mean=" & TEXT({delta_mean_cell},"+0.0;-0.0") & " fps, Δ SD=" & TEXT({delta_sd_cell},"+0.0;-0.0") & " fps",'
        f'"➤ SIGNIFICANT AND MEANINGFUL: The modification had a real, measurable effect. Δ Mean=" & TEXT({delta_mean_cell},"+0.0;-0.0") & " fps, Δ SD=" & TEXT({delta_sd_cell},"+0.0;-0.0") & " fps. Consider whether this improvement/degradation justifies the change.")))'
    )

    ws.cell(row=interp_row, column=stats_col, value=verdict_formula)
    ws.cell(row=interp_row, column=stats_col).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[interp_row].height = 80

    # Set column widths
    ws.column_dimensions[get_column_letter(stats_col)].width = 25
    ws.column_dimensions[get_column_letter(stats_col + 1)].width = 15
    ws.column_dimensions[get_column_letter(stats_col + 2)].width = 15

    # Charts note
    chart_note_row = interp_row + 5
    ws.cell(row=chart_note_row, column=stats_col,
           value="📊 Charts would show:")
    set_cell_style(ws.cell(row=chart_note_row, column=stats_col), font_bold=True, bg_color=COLOR_CALC_BG)
    ws.cell(row=chart_note_row + 1, column=stats_col, value="1. Overlapping histograms (Before=gray, After=green)")
    ws.cell(row=chart_note_row + 2, column=stats_col, value="2. Side-by-side box plots")
    ws.cell(row=chart_note_row + 3, column=stats_col, value="3. Sequential velocity plot through both conditions")
    ws.cell(row=chart_note_row + 4, column=stats_col, value="4. Delta bar chart (change in mean and SD)")

    # Protect sheet
    ws.protection.sheet = True
    ws.protection.password = 'reloading'

    # Unlock data entry cells
    for row in range(data_start_row, data_end_row + 1):
        for col in [2, 3]:
            ws.cell(row=row, column=col).protection = openpyxl.styles.Protection(locked=False)


def main():
    """Generate both blank and example Excel templates."""
    print("Generating Excel templates...")

    # Create blank version
    print("\n1. Creating blank template workbook...")
    wb_blank = create_workbook()
    create_readme_sheet(wb_blank)
    create_template_a(wb_blank, with_example_data=False)
    create_template_b(wb_blank, with_example_data=False)
    create_template_c(wb_blank, with_example_data=False)

    # Hide calculations sheet
    wb_blank["_Calculations"].sheet_state = 'hidden'

    wb_blank.save('Reloading_Analysis_Templates.xlsx')
    print("   ✓ Saved: Reloading_Analysis_Templates.xlsx")

    # Create example version
    print("\n2. Creating example template workbook...")
    wb_example = create_workbook()
    create_readme_sheet(wb_example)
    create_template_a(wb_example, with_example_data=True)
    create_template_b(wb_example, with_example_data=True)
    create_template_c(wb_example, with_example_data=True)

    wb_example["_Calculations"].sheet_state = 'hidden'

    wb_example.save('Reloading_Analysis_Templates_Examples.xlsx')
    print("   ✓ Saved: Reloading_Analysis_Templates_Examples.xlsx")

    print("\n✅ Excel templates generated successfully!")
    print("\nNext steps:")
    print("1. Open the files in Excel to review")
    print("2. Test with your own data")
    print("3. Note: Charts are noted in templates but must be created manually")
    print("   (openpyxl has limited chart support - manual creation gives better results)")
    print("\nPassword to unprotect sheets (if needed): reloading")


if __name__ == "__main__":
    main()
